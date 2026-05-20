"""
全球户外采购客户采集系统 - GitHub Actions 云端版 v2
双推送通道: QQ邮箱 + 微信(Server酱)
部署: 上传本文件 + .github/workflows/crawler.yml 到GitHub
"""

import os
import re
import time
import random
import sqlite3
import logging
import smtplib
import requests
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ===================== 日志 =====================
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", handlers=[logging.StreamHandler()])
logger = logging.getLogger(__name__)

# ===================== 配置 =====================
@dataclass(frozen=True)
class AppConfig:
    pass_score: int = 18
    high_value_score: int = 25
    factory_deduct: int = -40
    platform_deduct: int = -30
    max_pages: int = 5
    batch_size: int = 50
    db_path: str = "customer_final.db"
    output_dir: str = "客户数据"
    email_re: re.Pattern = field(default_factory=lambda: re.compile(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b"))
    phone_re: re.Pattern = field(default_factory=lambda: re.compile(r"\+?\d{1,3}[-\s]?\(?\d{2,4}\)?[-\s]?\d{3,4}[-\s]?\d{3,4}\b"))
    whatsapp_re: re.Pattern = field(default_factory=lambda: re.compile(r"(whatsapp|wa|whats app)[:\s]*(\+?\d{8,20})", re.IGNORECASE))
    product_words: Tuple[str, ...] = ("tent", "sleeping bag", "camping", "outdoor", "camp", "露营", "帐篷", "睡袋")
    buyer_words: Tuple[str, ...] = ("importer", "wholesaler", "distributor", "buyer", "retailer", "贸易", "采购")
    factory_words: Tuple[str, ...] = ("factory", "manufacturer", "oem", "odm", "producer", "生产", "工厂", "代工", "制造厂")
    bad_words: Tuple[str, ...] = ("alibaba", "amazon", "made-in-china", "aliexpress", "淘宝", "京东", "1688")
    contact_words: Tuple[str, ...] = ("owner", "founder", "ceo", "general manager", "purchasing manager", "buyer", "director", "采购", "负责人", "老板")
    contact_page_words: Tuple[str, ...] = ("contact", "about", "about us", "联系我们", "联系方式", "关于我们")

# ===================== 数据模型 =====================
@dataclass
class Customer:
    company: str
    url: Optional[str] = None
    country: str = ""
    email: Optional[str] = None
    phone: Optional[str] = None
    whatsapp: Optional[str] = None
    contact: Optional[str] = None
    score: int = 0
    page: int = 0
    lang: str = "英语"
    desc: Optional[str] = None
    created: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

# ===================== 精简国家列表 =====================
COUNTRY_SEARCH_KEY: Dict[str, List[str]] = {
    "美国": ["tent sleeping bag importer USA"],
    "英国": ["tent sleeping bag importer UK"],
    "德国": ["tent sleeping bag importer Germany"],
    "法国": ["tent sleeping bag importer France"],
    "意大利": ["tent sleeping bag importer Italy"],
    "西班牙": ["tent sleeping bag importer Spain"],
    "荷兰": ["tent sleeping bag importer Netherlands"],
    "加拿大": ["tent sleeping bag importer Canada"],
    "澳大利亚": ["tent sleeping bag importer Australia"],
    "日本": ["tent sleeping bag importer Japan"],
    "韩国": ["tent sleeping bag importer South Korea"],
}

# ===================== 数据库 =====================
class CustomerRepository:
    def __init__(self, config: AppConfig):
        self.config = config
        self._init_db()

    def _get_conn(self):
        conn = sqlite3.connect(self.config.db_path, isolation_level=None)
        conn.execute("PRAGMA journal_mode=WAL")
        return conn

    def _init_db(self):
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS customer (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company TEXT NOT NULL,
                url TEXT, country TEXT, email TEXT, phone TEXT, whatsapp TEXT, contact TEXT,
                score INTEGER DEFAULT 0, page INTEGER, lang TEXT, desc TEXT, created TEXT,
                UNIQUE(company)
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_customer_score ON customer(score DESC)")
        conn.close()

    def exists(self, company: str) -> bool:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM customer WHERE company=? LIMIT 1", (company,))
        result = cur.fetchone() is not None
        conn.close()
        return result

    def batch_insert(self, customers: List[Customer]):
        if not customers:
            return
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("BEGIN IMMEDIATE")
        try:
            cur.executemany(
                "INSERT OR IGNORE INTO customer (company, url, country, email, phone, whatsapp, contact, score, page, lang, desc, created) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                [(c.company, c.url, c.country, c.email, c.phone, c.whatsapp, c.contact, c.score, c.page, c.lang, c.desc, c.created) for c in customers]
            )
            cur.execute("COMMIT")
            logger.info(f"写入 {len(customers)} 条")
        except Exception as e:
            cur.execute("ROLLBACK")
            logger.error(f"写入失败: {e}")
        finally:
            conn.close()

    def fetch_all(self) -> List[tuple]:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("SELECT company, country, url, email, phone, whatsapp, contact, score, page, lang, desc, created FROM customer ORDER BY score DESC")
        rows = cur.fetchall()
        conn.close()
        return rows

    def count(self) -> int:
        conn = self._get_conn()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM customer")
        count = cur.fetchone()[0]
        conn.close()
        return count

# ===================== 评分器 =====================
class Scorer:
    def score(self, company: str, desc: str, url: str, country: str, config: AppConfig) -> int:
        text = f"{company} {desc}".lower()
        score = 0
        for w in config.product_words:
            if w.lower() in text:
                score += 6
        score = min(score, 30)
        for w in config.buyer_words:
            if w.lower() in text:
                score += 7
        if country.lower() in text:
            score += 15
        if url and len(url) > 10:
            score += 8
        if config.email_re.search(desc) or config.phone_re.search(desc):
            score += 4
        for w in config.factory_words:
            if w.lower() in text:
                score += config.factory_deduct
                break
        for w in config.bad_words:
            if w.lower() in f"{url} {desc}".lower():
                score += config.platform_deduct
                break
        return score

# ===================== 爬虫 =====================
class GoogleCrawler:
    def __init__(self, config: AppConfig):
        self.config = config
        self.driver = None

    def __enter__(self):
        self._start()
        return self

    def __exit__(self, *args):
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass

    def _start(self):
        options = Options()
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--disable-images")
        try:
            self.driver = webdriver.Chrome(options=options)
            logger.info("Chrome启动成功")
        except:
            try:
                from webdriver_manager.chrome import ChromeDriverManager
                service = Service(ChromeDriverManager().install())
                self.driver = webdriver.Chrome(service=service, options=options)
                logger.info("Chrome启动成功(自动下载)")
            except Exception as e:
                logger.error(f"Chrome启动失败: {e}")
                raise
        self.driver.set_page_load_timeout(30)
        self.driver.get("https://www.google.com")
        logger.info("已访问Google")

    def search(self, keyword: str, page: int) -> List[Dict]:
        url = f"https://www.google.com/search?q={keyword}&start={(page-1)*10}&hl=en"
        try:
            self.driver.get(url)
            time.sleep(random.uniform(5, 10))
        except:
            return []
        soup = BeautifulSoup(self.driver.page_source, "lxml")
        container = soup.find("div", id="search")
        if not container:
            return []
        items = container.find_all("div", attrs={"data-result-index": True})
        if not items:
            items = container.find_all("div", class_=re.compile(r"^(g|MjjYud|kvH3mc)$"))
        results = []
        for item in items:
            try:
                title = item.find("h3")
                if not title:
                    continue
                company = title.get_text(strip=True)
                link = item.find("a")
                url = ""
                if link and link.get("href"):
                    href = link["href"]
                    if href.startswith("/url?q="):
                        url = href.split("&")[0].replace("/url?q=", "")
                    elif href.startswith("http"):
                        url = href
                desc = item.get_text(strip=True)[:400]
                results.append({"company": company, "url": url, "desc": desc})
            except:
                continue
        return results

    def extract_contact(self, html: str) -> Dict[str, Optional[str]]:
        result = {"email": None, "phone": None, "whatsapp": None, "contact": None}
        try:
            emails = self.config.email_re.findall(html)
            if emails:
                result["email"] = emails[0].lower()
            phones = self.config.phone_re.findall(html)
            if phones:
                result["phone"] = phones[0].strip()
            wa = self.config.whatsapp_re.findall(html)
            if wa:
                result["whatsapp"] = wa[0][1].strip()
            elif result["phone"] and result["phone"].startswith("+"):
                result["whatsapp"] = result["phone"]
            text = html.lower()
            for word in self.config.contact_words:
                if word.lower() in text:
                    result["contact"] = word.title()
                    break
        except:
            pass
        return result

# ===================== 双通道推送 =====================
class Notifier:
    def __init__(self):
        self.email = os.environ.get("PUSH_EMAIL", "")
        self.smtp_password = os.environ.get("SMTP_PASSWORD", "")
        self.wx_key = os.environ.get("PUSH_WX_KEY", "")

    def send_email(self, subject: str, body: str, attachment_path: str = ""):
        if not self.email or not self.smtp_password:
            logger.warning("未配置邮箱，跳过邮件推送")
            return False
        try:
            msg = MIMEMultipart()
            msg["From"] = self.email
            msg["To"] = self.email
            msg["Subject"] = subject
            msg.attach(MIMEText(body, "plain", "utf-8"))
            if attachment_path and os.path.exists(attachment_path):
                with open(attachment_path, "rb") as f:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(attachment_path)}")
                msg.attach(part)
            server = smtplib.SMTP_SSL("smtp.qq.com", 465)
            server.login(self.email, self.smtp_password)
            server.sendmail(self.email, self.email, msg.as_string())
            server.quit()
            logger.info("✅ 邮件发送成功")
            return True
        except Exception as e:
            logger.error(f"邮件发送失败: {e}")
            return False

    def send_wechat(self, title: str, content: str):
        if not self.wx_key:
            logger.warning("未配置微信Key，跳过微信推送")
            return False
        try:
            url = f"https://sctapi.ftqq.com/{self.wx_key}.send"
            data = {"title": title, "desp": content}
            response = requests.post(url, data=data, timeout=10)
            if response.json().get("code") == 0:
                logger.info("✅ 微信推送成功")
                return True
            else:
                logger.error(f"微信推送失败: {response.text}")
                return False
        except Exception as e:
            logger.error(f"微信推送异常: {e}")
            return False

    def notify(self, subject: str, body: str, attachment_path: str = ""):
        self.send_email(subject, body, attachment_path)
        self.send_wechat(subject, body)

# ===================== 导出Excel =====================
def export_excel(repo: CustomerRepository, output_dir: str) -> str:
    os.makedirs(output_dir, exist_ok=True)
    path = os.path.join(output_dir, "全球户外采购客户(云端版).xlsx")
    rows = repo.fetch_all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "客户列表"
    headers = ["公司名称", "国家", "官网", "邮箱", "电话", "WhatsApp", "联系人", "评分", "页码", "语言", "简介", "时间"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E74B5", end_color="2E74B5", fill_type="solid")
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    wb.save(path)
    logger.info(f"Excel导出: {path} ({len(rows)}条)")
    return path

# ===================== 主程序 =====================
def main():
    config = AppConfig()
    repo = CustomerRepository(config)
    scorer = Scorer()
    notifier = Notifier()

    print("=" * 60)
    print("全球户外采购客户采集系统 - GitHub Actions云端版 v2")
    print("推送通道: QQ邮箱 + 微信(Server酱)")
    print("=" * 60)
    print(f"当前客户库: {repo.count()} 家")
    print("=" * 60)

    countries = list(COUNTRY_SEARCH_KEY.items())[:3]
    total_new = 0

    with GoogleCrawler(config) as crawler:
        for country, keywords in countries:
            for keyword in keywords:
                logger.info(f"\n🌍 采集: {country} | {keyword}")
                for page in range(1, config.max_pages + 1):
                    logger.info(f"  第{page}页...")
                    try:
                        results = crawler.search(keyword, page)
                    except Exception as e:
                        logger.error(f"搜索失败: {e}")
                        break

                    page_valid = 0
                    buffer = []
                    for raw in results:
                        company = raw.get("company", "未知")
                        url = raw.get("url", "")
                        desc = raw.get("desc", "")
                        if repo.exists(company):
                            continue
                        score = scorer.score(company, desc, url, country, config)
                        if score < config.pass_score:
                            continue
                        contact = crawler.extract_contact(crawler.driver.page_source)
                        customer = Customer(
                            company=company, url=url, country=country,
                            email=contact["email"], phone=contact["phone"],
                            whatsapp=contact["whatsapp"], contact=contact["contact"],
                            score=score, page=page, desc=desc
                        )
                        buffer.append(customer)
                        page_valid += 1

                    if buffer:
                        repo.batch_insert(buffer)
                        total_new += len(buffer)

                    logger.info(f"  第{page}页: {page_valid}家")
                    time.sleep(random.uniform(10, 15))

    excel_path = export_excel(repo, config.output_dir)

    subject = f"【获客完成】新增{total_new}家，累计{repo.count()}家客户"
    body = f"""
📊 采集时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}
📈 本次新增: {total_new} 家
📊 累计客户: {repo.count()} 家

💡 说明:
- Excel附件为完整客户列表，按评分从高到低排序
- 建议优先联系评分25分以上的客户
- 每天自动运行，无需手动操作

📁 文件: 全球户外采购客户(云端版).xlsx
    """

    notifier.notify(subject, body, excel_path)

    print("\n" + "=" * 60)
    print(f"✅ 完成！新增 {total_new} 家，累计 {repo.count()} 家")
    print("=" * 60)

if __name__ == "__main__":
    main()
